#!/usr/bin/env python

import datetime
from os.path import splitext

from openpyxl.reader.excel import load_workbook
import six

from csvkit import CSVKitWriter 
from csvkit.typeinference import NULL_TIME

def normalize_datetime(dt):
    if dt.microsecond == 0:
        return dt

    ms = dt.microsecond

    if ms < 1000:
        return dt.replace(microsecond=0)
    elif ms > 999000:
        return dt.replace(microsecond=0) + datetime.timedelta(seconds=1)

    return dt

def has_date_elements(cell):
    """
    Try to use formatting to determine if a cell contains only time info.

    See: http://office.microsoft.com/en-us/excel-help/number-format-codes-HP005198679.aspx
    """
    if 'd' in cell.number_format or \
        'y' in cell.number_format:

        return True

    return False

def xlsx2csv(f, output=None, **kwargs):
    """
    Convert an Excel .xlsx file to csv.

    Note: Unlike other convertor's, this one allows output columns to contain mixed data types.
    Blank headers are also possible.
    """
    streaming = True if output else False

    if not streaming:
        output = six.StringIO()

    writer = CSVKitWriter(output)

    book = load_workbook(f, use_iterators=True, data_only=True)
    base, ext = splitext(f.name)

    if 'sheet' in kwargs:
        sheet_to_operate_on = book.get_sheet_by_name(kwargs['sheet'])
    else:
        sheet_to_operate_on = book.get_active_sheet()

    if kwargs.get('write_all_sheets'):
        sheets = book.worksheets
    else:
        sheets = [sheet_to_operate_on]

    for sheet_index, sheet in enumerate(sheets):
        if kwargs.get('write_all_sheets'):
            write_rows = []

        for i, row in enumerate(sheet.iter_rows()):
            if i == 0:
                first_row = [c.value for c in row]
                if sheet == sheet_to_operate_on:
                    writer.writerow(first_row)
                if kwargs.get('write_all_sheets'):
                    write_rows.append(first_row)
                continue

            out_row = []

            for c in row:
                value = c.value

                if value.__class__ is datetime.datetime:
                    # Handle default XLSX date as 00:00 time 
                    if value.date() == datetime.date(1904, 1, 1) and not has_date_elements(c):
                        value = value.time() 

                        value = normalize_datetime(value)
                    elif value.time() == NULL_TIME:
                        value = value.date()
                    else:
                        value = normalize_datetime(value)
                elif value.__class__ is float:
                    if value % 1 == 0:
                        value = int(value)

                if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                    value = value.isoformat()

                out_row.append(value)

            if sheet == sheet_to_operate_on:
                writer.writerow(out_row)
            if kwargs.get('write_all_sheets'):
                write_rows.append(out_row)

        if kwargs.get('write_all_sheets'):
            with open('%s_%d.csv' % (base, sheet_index), 'wb') as f:
                CSVKitWriter(f).writerows(write_rows)

    if not streaming:
        data = output.getvalue()
        return data

    # Return empty string when streaming
    return ''

